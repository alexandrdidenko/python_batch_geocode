import multiprocessing

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from tqdm import tqdm


# ------------------------------------- Работа с Excel -----------------------------------------------------------------
class ExcelWorkbook:
    """Создать новый файл excel"""

    def __init__(self, filename):
        self.filename = filename

    def create(self):
        """Создать файл excel"""
        print(ExcelWorkbook.create.__doc__)
        wb = openpyxl.Workbook()
        return wb

    def save(self, wb):
        """Сохранить файл excel"""
        print(ExcelWorkbook.save.__doc__)
        wb.save(filename=self.filename)

    def read(self):
        """Відкриваємо файл та зчитуємо дані в словник"""
        print(ExcelWorkbook.read.__doc__)
        excel_file = openpyxl.load_workbook(self.filename)
        employees_sheet = excel_file.active
        res = []

        for mr in range(1, employees_sheet.max_row + 1):
            rr = []
            for x in range(1, employees_sheet.max_column + 1):
                r = employees_sheet.cell(row=mr, column=x).value
                rr.append(r)
            res.append(rr)
        return res


class ExcelSheet:
    """Создаеть новый лист excel"""

    def __init__(self, all_rows, columns, num=(0,), dec=(0,), dat=(0,), alig_center=(0,)):
        """конструктор"""
        self.all_rows = all_rows
        self.columns = columns
        self.num = num
        self.dec = dec
        self.dat = dat
        self.alig_center = alig_center

    def sheet_active(self, wb, title='data'):
        """Aктивировать первый лист"""
        print(ExcelSheet.sheet_active.__doc__)
        ws = wb.active
        ws.title = title
        return ws

    def sheet_add(self, wb, title='data'):
        """Создать новый лист"""
        print(ExcelSheet.sheet_add.__doc__)
        # ws = wb.active
        # ws.title = title
        ws = wb.create_sheet(title)
        return ws

    def write(self, ws):
        """Записать данные"""
        print(ExcelSheet.write.__doc__)
        # шрифт первой строки
        font_first = Font(name='Century Gothic',
                          size=10,
                          bold=True,
                          italic=False,
                          vertAlign=None,
                          underline='none',
                          strike=False,
                          color='FF000000')
        # шрифт всего текста

        font = Font(name='Century Gothic',
                    size=10,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

        # цвет первой строки
        fill_first = PatternFill("solid", fgColor="A6A6A6")

        # границы ячеек
        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style=None,
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style=None,
                                     color='FF000000'),
                        vertical=Side(border_style=None,
                                      color='FF000000'),
                        horizontal=Side(border_style=None,
                                        color='FF000000'))

        # выравнивание первой строки
        alignment_first = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=True,
                                    shrink_to_fit=True,
                                    indent=0)

        # выравнивание
        alignment = Alignment(horizontal='left',
                              vertical='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)

        # выравнивание по центру в строках
        alignment_center = Alignment(horizontal='center',
                                     vertical='center',
                                     text_rotation=0,
                                     wrap_text=False,
                                     shrink_to_fit=False,
                                     indent=0)

        # заполняем шапку
        c = 1
        for col in self.columns:
            ws.cell(column=c, row=1, value=col)
            ws.cell(column=c, row=1).font = font_first
            ws.cell(column=c, row=1).fill = fill_first
            ws.cell(column=c, row=1).border = border
            ws.cell(column=c, row=1).alignment = alignment_first
            c += 1

        # пишем заполняем файл
        r = 2
        len_rows = len(self.all_rows)

        for i in tqdm(range(len_rows), colour='green'):  # цикл который рисует прогресбар
            row = self.all_rows[i]
            # for row in self.all_rows:  # старый цикл до того как сделал прогрес бар
            #     print(str(r) + ' из ' + str(len_rows))
            c = 1
            for col in row:
                if c in self.num:
                    if col is None or col == '':
                        pass
                    else:
                        ws.cell(column=c, row=r, value=int(col)).number_format = '0'
                elif c in self.dec:
                    if col is None:
                        pass
                    else:
                        try:
                            ws.cell(column=c, row=r, value=float(col)).number_format = '0.0000'
                        except ValueError:
                            ws.cell(column=c, row=r, value=col).number_format = '0.0000'
                elif c in self.dat:
                    ws.cell(column=c, row=r, value=col).number_format = 'DD-MM-YYYY'
                else:
                    ws.cell(column=c, row=r, value=col).number_format = '@'
                ws.cell(column=c, row=r).font = font
                ws.cell(column=c, row=r).border = border
                if c in self.alig_center:
                    ws.cell(column=c, row=r).alignment = alignment_center
                else:
                    ws.cell(column=c, row=r).alignment = alignment
                c = c + 1
            r = r + 1

        # столбики по ширине текста
        MIN_WIDTH = 20
        MAX_WIDTH = 70
        for i, column_cells in enumerate(ws.columns, start=1):
            width = (
                length
                if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                                  for cell in column_cells)) >= MIN_WIDTH
                else MIN_WIDTH
            )
            if width >= MAX_WIDTH:
                ws.column_dimensions[get_column_letter(i)].width = MAX_WIDTH
            else:
                ws.column_dimensions[get_column_letter(i)].width = width
